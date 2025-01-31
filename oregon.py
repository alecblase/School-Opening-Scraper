import logging

import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import pandas as pd
from datetime import date
from datetime import datetime


def main():
    #logging.basicConfig(filename='app.log', filemode='a', format='%(asctime)s - %(message)s', level=logging.INFO)
    download_xslx()
    #logging.info("Received Oregon Data", exc_info=False);
    copy_to_new_csv()
    #logging.info("Wrote Oregon Data", exc_info=False);

def download_xslx():
    # Get html of page
    url = "https://www.oregon.gov/ode/students-and-family/healthsafety/Pages/2020-21-School-Status.aspx"
    html = requests.get(url).content
    soup = BeautifulSoup(html, 'html.parser')

    # Get the most recent update link
    path = soup.body.main.div.contents[9].li.a['href']
    dataUrl = "https://www.oregon.gov/" + path

    # Retrieve cvs file
    originalFile = requests.get(dataUrl)
    open('temp/OregonOriginal.xlsx', 'wb').write(originalFile.content)

def copy_to_new_csv():
    wb = load_workbook('temp/OregonOriginal.xlsx')
    districtSheet = wb['District List']
    prevDistrict = ""
    inputRow = 0
    districtIndex = -1
    df = pd.DataFrame(
        columns=['district', 'on-site school count', 'hybrid school count',
                 'distance school count', 'distance w/LIPI school count', 'report date', 'date scraped'])

    for row in districtSheet.iter_rows(values_only=True):
        if inputRow == 0:  # Skip column headers
            inputRow += 1
            continue
        curDistrict = row[2]
        curMode = row[5]
        if curDistrict is None:  # End of input file
            break

        onSite = 0
        hybrid = 0
        distance = 0
        distanceLIPI = 0
        if curMode == "On-Site":
            onSite = 1
        elif curMode == "Hybrid":
            hybrid = 1
        elif curMode == "Comprehensive Distance Learning":
            distance = 1
        elif curMode == "Comprehensive Distance Learning w/LIPI":
            distanceLIPI = 1
        if curDistrict == prevDistrict:  # Continue with previous district
            if onSite == 1:
                df.iat[districtIndex, 1] = df.iat[districtIndex, 1] + 1
            elif hybrid == 1:
                df.iat[districtIndex, 2] = df.iat[districtIndex, 2] + 1
            elif distance == 1:
                df.iat[districtIndex, 3] = df.iat[districtIndex, 3] + 1
            elif distanceLIPI == 1:
                df.iat[districtIndex, 4] = df.iat[districtIndex, 4] + 1
        else:  # Start new district
            districtIndex += 1
            reportWeek = row[3]
            dateUpdated = reportWeek[12:]  # End date of report week

            newDistrictRow = pd.Series(data={'district': curDistrict, 'on-site school count': onSite,
                                             'hybrid school count': hybrid, 'distance school count': distance,
                                             'distance w/LIPI school count': distanceLIPI, 'report date': dateUpdated,
                                             'date scraped': date.today()})

            df = df.append(newDistrictRow, ignore_index=True)

        inputRow += 1  # End for
        prevDistrict = curDistrict
    df.to_csv('out/OR_' + datetime.now().strftime('%Y%m%d') + '.csv', index=False)  # Copy dataframe to CSV

#main()
